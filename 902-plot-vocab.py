import pandas as pd
from matplotlib import pyplot as plt
import numpy as np
import sys
import os
from openpyxl import Workbook,load_workbook
import matplotlib as mpl
from matplotlib.font_manager import FontProperties
#mpl.rcParams['font.sans-serif'] = ['simkai','SimHei']  # 汉字字体,优先使用楷体，如果找不到楷体，则使用黑体
#mpl.rcParams['font.size'] = 12  # 字体大小
#mpl.rcParams['axes.unicode_minus'] = False  # 正常显示负号
linestyles = ['o-', 's-', '*-','>-','<-','g-', '.-', 'o-', 's-', '*-']
colours = [
    'r',
    'b',
    'g',
    'y',
    'c',
    'orange',
    'purple',
    'pink',
    'gray',
    'black',
    ]

def run():
    wb=load_workbook('./90-vocab.xlsx')
    sheets=['story','SPP']
    #colors=['gold','deepskyblue','r','lime']
    # colors=['#663300','royalblue','#FF9933','r']
    colors=['#663300','royalblue','#FF9933','#CC3333']
    # labels=['14.67% Animals','12.70% Interpersonal Relationships','46.84% Concrete Nouns','26.00% Abstract Nouns']
    labels=['21.65% Animals','23.53% Interpersonal Relationships','33.58% Concrete Nouns','21.24% Abstract Nouns']
    xlabels=['noun words of en-story','noun words of en-moral','noun words of zh-story','noun words of zh-moral']
    zhfont = FontProperties(fname=r"./simhei.ttf")
    print('labels', labels)
    for sheet in sheets:
        print('sheet',sheet)
        
        #mngr = plt.get_current_fig_manager()  # 获取当前figure manager
        #mngr.window.wm_geometry("+0-100")
        plt.subplots(1,1,figsize=(14,4))
        #plt.figure(figsize=(14, 4))
        
        ws=wb[sheet]
        items=[[[],[]], [[],[]], [[],[]], [[],[]]]
        ticklabel=[]
        #width=50/50
        width=0.95
        flag=0
        for i in range(2,52):
            name =ws['A%d'%(i)].value
            freq =int(ws['B%d'%(i)].value)
            #print(name,freq)
            label=int(ws['C%d'%(i)].value)
            trans=ws['D%d'%(i)].value
            if trans is not None:
                name+='('+trans+')'
                flag=1
            items[label-1][0].append(i-2)
            items[label-1][1].append(freq)
            ticklabel.append(name)
        for i in range(4):
            #print(items[i][0],items[i][2])
            if len(items[i][0])==0:
                items[i][0].append(0)
                items[i][1].append(1)
            plt.rcParams['legend.fontsize']=12
            plt.bar(items[i][0],items[i][1],color=colors[i],align='center',label=labels[i],width=width)
            plt.xticks(range(50),ticklabel,fontsize=12.5,rotation='80', fontproperties=zhfont)
        plt.legend(fontsize=16)
        print("1234")
        #ax = plt.gca()
        #plt.axis('auto')
        if not flag:
            plt.gcf().subplots_adjust(bottom=0.302)
        else:
            plt.gcf().subplots_adjust(bottom=0.4)
        #plt.xlabel('noun words of zh-story')
        plt.ylabel('Frequency', fontsize=16)
        plt.show()
        #print('st')
        plt.savefig('./%s.pdf'%(sheet))
        plt.close()
        #break

if __name__=='__main__':
    print(mpl.matplotlib_fname())
    run()
